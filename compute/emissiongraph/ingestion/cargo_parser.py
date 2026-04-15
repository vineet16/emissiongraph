"""Parser for the 'Cargo Handled' sheet — monthly cargo throughput in MT.

Actual layout (vertical date-value):
  Row 1: Port ID in B1
  Row 3: "Cargo Handled" label
  Row 4: "Month" | "Cargo (Metric Tons)" headers
  Row 5: FY label (e.g. "FY24-25")
  Rows 6-17: datetime (2024-04-01) | numeric value
  Row 18: "Total" | total value
"""

from __future__ import annotations

from datetime import datetime

from openpyxl.worksheet.worksheet import Worksheet

from emissiongraph.facts.schema import CellRef, Measurement

# Full month names (kept for backward compat with other modules that import these)
MONTH_NAMES = [
    "april", "may", "june", "july", "august", "september",
    "october", "november", "december", "january", "february", "march",
]

# FY month mapping
FY_MONTH_MAP = {
    "april": "04", "may": "05", "june": "06", "july": "07",
    "august": "08", "september": "09", "october": "10",
    "november": "11", "december": "12", "january": "01",
    "february": "02", "march": "03",
}


def _fy_to_start_year(fy: str) -> int:
    """'FY24-25' -> 2024, 'FY23-24' -> 2023."""
    parts = fy.replace("FY", "").split("-")
    y = int(parts[0])
    return 2000 + y if y < 100 else y


def _month_to_period_value(month_name: str, fy: str) -> str:
    """Convert month name + FY to period value like '2024-04'."""
    start_year = _fy_to_start_year(fy)
    mm = FY_MONTH_MAP.get(month_name.lower().strip())
    if mm is None:
        return ""
    month_int = int(mm)
    year = start_year if month_int >= 4 else start_year + 1
    return f"{year}-{mm}"


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel column letter."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _datetime_to_period_value(dt: datetime) -> str:
    """Convert a datetime to period value like '2024-04'."""
    return f"{dt.year}-{dt.month:02d}"


def parse_cargo_sheet(
    ws: Worksheet,
    port_id: str,
    fy: str,
    workbook_name: str,
) -> list[Measurement]:
    """Extract monthly cargo throughput from the Cargo Handled sheet.

    Format: vertical layout with dates in col B and values in col C.
    """
    measurements: list[Measurement] = []

    # Find the header row with "Month" label
    header_row = None
    date_col = None
    value_col = None

    for row_idx in range(1, min((ws.max_row or 20) + 1, 20)):
        for col_idx in range(1, min((ws.max_column or 20) + 1, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str) and val.strip().lower() == "month":
                header_row = row_idx
                date_col = col_idx
                break
        if header_row:
            break

    if not header_row or not date_col:
        return measurements

    # Find the first data row (row with a datetime in the date column)
    first_data_row = None
    for row_idx in range(header_row + 1, min((ws.max_row or 30) + 1, 30)):
        cell = ws.cell(row=row_idx, column=date_col).value
        if isinstance(cell, datetime):
            first_data_row = row_idx
            break
        if isinstance(cell, str):
            try:
                datetime.fromisoformat(cell.split(" ")[0])
                first_data_row = row_idx
                break
            except (ValueError, IndexError):
                continue

    if not first_data_row:
        return measurements

    # Find the value column by scanning all columns in the first data row
    # for the one with a numeric value (handles misaligned headers like P10)
    for vc in range(date_col + 1, (ws.max_column or 20) + 1):
        val = ws.cell(row=first_data_row, column=vc).value
        if val is not None:
            try:
                float(val)
                value_col = vc
                break
            except (ValueError, TypeError):
                continue

    if not value_col:
        return measurements

    # Parse data rows below the FY label row
    for row_idx in range(header_row + 1, (ws.max_row or 50) + 1):
        date_cell = ws.cell(row=row_idx, column=date_col).value
        val_cell = ws.cell(row=row_idx, column=value_col).value

        # Skip FY label row
        if isinstance(date_cell, str) and date_cell.strip().upper().startswith("FY"):
            continue

        # Stop at Total row
        if isinstance(date_cell, str) and "total" in date_cell.strip().lower():
            break

        # Parse datetime
        if isinstance(date_cell, datetime):
            period_value = _datetime_to_period_value(date_cell)
        elif isinstance(date_cell, str):
            # Try parsing string datetime
            try:
                dt = datetime.fromisoformat(date_cell.split(" ")[0])
                period_value = _datetime_to_period_value(dt)
            except (ValueError, IndexError):
                continue
        else:
            continue

        # Parse value
        qty = 0.0
        if val_cell is not None:
            if isinstance(val_cell, str) and val_cell.strip() == "-":
                qty = 0.0
            else:
                try:
                    qty = float(val_cell)
                except (ValueError, TypeError):
                    continue

        measurements.append(
            Measurement(
                port_id=port_id,
                fy=fy,
                period="monthly",
                period_value=period_value,
                fuel_type="Cargo",
                sub_type=None,
                measure="consumption",
                quantity=qty,
                unit="MT",
                source_cell=CellRef(
                    workbook=workbook_name,
                    sheet="Cargo Handled",
                    cell=f"{_col_letter(value_col)}{row_idx}",
                    row=row_idx,
                    col=value_col,
                ),
            )
        )

    return measurements
