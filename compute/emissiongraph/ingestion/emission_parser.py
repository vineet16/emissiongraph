"""
EmissionGraph — Port Workbook Parser (v0.2)

Strategy:
1. Anchor on the unique "GHG Intensity" row in 305-4 — present in every workbook.
2. Grand total emissions = the cell directly above the intensity row.
3. Scope 1 components (diesel-stat, diesel-mob, petrol, hfhsd_ifo, other_fuels)
   are read by section + "GHG Emission (tCO2e)" or "Total Emissions (tCO2)" label
   within that section's columns. Sections are identified by row 3 banner text.
4. Scope 2 electricity = grand_total - sum(scope_1_components). This is robust
   to single-source vs multi-source electricity layouts because we never need
   to find every electricity sub-cell.
5. Cargo total from "Cargo Handled" sheet via "Total" row label.
"""

from dataclasses import dataclass, asdict
from typing import Optional
import openpyxl
import json


@dataclass
class CellRef:
    workbook: str
    sheet: str
    cell: str
    row: int
    col: int


@dataclass
class HeadlineMetrics:
    port_id: str
    fy: str
    cargo_mt: float
    total_emissions_tco2e: float
    scope2_electricity_tco2e: float        # derived: grand - sum(scope1)
    scope1_diesel_stationary_tco2e: float
    scope1_diesel_mobile_tco2e: float
    scope1_petrol_tco2e: float
    scope1_hfhsd_ifo_tco2e: float
    scope1_other_fuels_tco2e: float
    ghg_intensity_kg_per_mt: float
    cell_refs: dict


SECTION_KEYWORDS = {
    "electricity": ["electricity"],
    "diesel": ["diesel"],
    "petrol": ["petrol"],
    "hfhsd_ifo": ["hfhsd", "ifo"],
    "other_fuels": ["other fuel"],
}


def _col_letter(col_idx: int) -> str:
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _addr(r, c): return f"{_col_letter(c)}{r}"


def _cref(wb, sh, r, c):
    return CellRef(workbook=wb, sheet=sh, cell=_addr(r, c), row=r, col=c)


def _find_section_columns(sheet) -> dict:
    """Section banner row is row 3 in display terms (1-indexed)."""
    section_starts = []
    for col in range(1, sheet.max_column + 1):
        v = sheet.cell(row=3, column=col).value
        if isinstance(v, str):
            vl = v.lower().strip()
            for name, kws in SECTION_KEYWORDS.items():
                if any(kw in vl for kw in kws):
                    section_starts.append((col, name))
                    break

    sections = {}
    for i, (start, name) in enumerate(section_starts):
        end = section_starts[i + 1][0] - 1 if i + 1 < len(section_starts) else sheet.max_column
        sections[name] = (start, end)
    return sections


def _find_label_row_in_cols(sheet, label_substring: str, col_range: tuple,
                              search_left_offset: int = 1) -> Optional[tuple]:
    """Find first row where any cell in the col range (slightly widened left) matches label.
    Returns (row, label_col) or None."""
    start, end = col_range
    search_start = max(1, start - search_left_offset)
    for row in range(1, sheet.max_row + 1):
        for col in range(search_start, end + 1):
            v = sheet.cell(row=row, column=col).value
            if isinstance(v, str) and label_substring.lower() in v.lower():
                return (row, col)
    return None


def _values_right_of(sheet, label_pos: tuple, col_range: tuple,
                       max_offset: int = 8) -> list:
    """Return list of (value, row, col) for numeric cells right of label within section."""
    label_row, label_col = label_pos
    _, end = col_range
    out = []
    for off in range(1, max_offset + 1):
        col = label_col + off
        if col > end + 2: break  # tolerate slight overshoot
        v = sheet.cell(row=label_row, column=col).value
        if isinstance(v, (int, float)) and v is not None:
            out.append((float(v), label_row, col))
    return out


def parse_cargo_total(workbook_path: str) -> tuple:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = wb["Cargo Handled"]
    for row in range(1, sheet.max_row + 1):
        v = sheet.cell(row=row, column=2).value
        if isinstance(v, str) and v.strip().lower() == "total":
            for col in range(3, sheet.max_column + 1):
                cv = sheet.cell(row=row, column=col).value
                if isinstance(cv, (int, float)) and cv > 0:
                    return float(cv), _cref(workbook_path, "Cargo Handled", row, col)
    raise ValueError(f"No cargo total in {workbook_path}")


def parse_305_4(workbook_path: str) -> dict:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = wb["305-4"]
    sections = _find_section_columns(sheet)

    result = {
        "scope1_diesel_stationary_tco2e": 0.0,
        "scope1_diesel_mobile_tco2e": 0.0,
        "scope1_petrol_tco2e": 0.0,
        "scope1_hfhsd_ifo_tco2e": 0.0,
        "scope1_other_fuels_tco2e": 0.0,
        "total_emissions_tco2e": 0.0,
        "ghg_intensity": 0.0,
        "cell_refs": {},
    }

    # Anchor: find GHG Intensity row (unique label, always in column 2)
    intensity_row = None
    for r in range(1, sheet.max_row + 1):
        v = sheet.cell(row=r, column=2).value
        if isinstance(v, str) and "ghg intensity" in v.lower():
            intensity_row = r
            break

    if intensity_row:
        intensity_val = sheet.cell(row=intensity_row, column=3).value
        if isinstance(intensity_val, (int, float)):
            result["ghg_intensity"] = float(intensity_val)
            result["cell_refs"]["ghg_intensity"] = _cref(
                workbook_path, "305-4", intensity_row, 3)

        # Grand total = directly above intensity
        for offset in [1, 2]:  # try 1 row above, then 2
            grand_row = intensity_row - offset
            grand_label = sheet.cell(row=grand_row, column=2).value
            grand_val = sheet.cell(row=grand_row, column=3).value
            if (isinstance(grand_label, str)
                    and "total emissions" in grand_label.lower()
                    and isinstance(grand_val, (int, float))):
                result["total_emissions_tco2e"] = float(grand_val)
                result["cell_refs"]["total_emissions"] = _cref(
                    workbook_path, "305-4", grand_row, 3)
                break

    # --- Diesel: GHG Emission (tCO2e) row in diesel section ---
    if "diesel" in sections:
        rng = sections["diesel"]
        lbl = _find_label_row_in_cols(sheet, "ghg emission (tco2e)", rng)
        if lbl:
            vals = _values_right_of(sheet, lbl, rng)
            if len(vals) >= 3:
                # stationary, mobile, total
                result["scope1_diesel_stationary_tco2e"] = vals[0][0]
                result["scope1_diesel_mobile_tco2e"] = vals[1][0]
                result["cell_refs"]["diesel_stationary"] = _cref(
                    workbook_path, "305-4", vals[0][1], vals[0][2])
                result["cell_refs"]["diesel_mobile"] = _cref(
                    workbook_path, "305-4", vals[1][1], vals[1][2])
            elif len(vals) == 1:
                has_stat = False; has_mob = False
                for c in range(rng[0], rng[1] + 1):
                    h = sheet.cell(row=4, column=c).value
                    if isinstance(h, str):
                        if "stationary" in h.lower(): has_stat = True
                        if "mobile" in h.lower(): has_mob = True
                if has_stat and has_mob:
                    result["scope1_diesel_mobile_tco2e"] = vals[0][0]
                else:
                    result["scope1_diesel_mobile_tco2e"] = vals[0][0]

    # --- Petrol ---
    if "petrol" in sections:
        rng = sections["petrol"]
        lbl = _find_label_row_in_cols(sheet, "ghg emission (tco2e)", rng)
        if lbl:
            vals = _values_right_of(sheet, lbl, rng)
            if vals:
                result["scope1_petrol_tco2e"] = vals[0][0]
                result["cell_refs"]["petrol"] = _cref(
                    workbook_path, "305-4", vals[0][1], vals[0][2])

    # --- HFHSD/IFO: stationary + mobile values, take rightmost (total) ---
    if "hfhsd_ifo" in sections:
        rng = sections["hfhsd_ifo"]
        lbl = _find_label_row_in_cols(sheet, "ghg emission (tco2e)", rng)
        if lbl:
            vals = _values_right_of(sheet, lbl, rng)
            if vals:
                v, vrow, vcol = vals[-1]
                result["scope1_hfhsd_ifo_tco2e"] = v
                result["cell_refs"]["hfhsd_ifo"] = _cref(
                    workbook_path, "305-4", vrow, vcol)

    # --- Other Fuels (annual: Acetylene, LPG, etc.) ---
    if "other_fuels" in sections:
        rng = sections["other_fuels"]
        for r in range(1, sheet.max_row + 1):
            for c in range(rng[0], min(rng[1] + 1, sheet.max_column + 1)):
                v = sheet.cell(row=r, column=c).value
                if isinstance(v, str) and "total emissions" in v.lower():
                    vals = _values_right_of(sheet, (r, c), rng, max_offset=4)
                    if vals:
                        vv, vr, vc = vals[-1]
                        result["scope1_other_fuels_tco2e"] = vv
                        result["cell_refs"]["other_fuels"] = _cref(
                            workbook_path, "305-4", vr, vc)
                    break
            if result["scope1_other_fuels_tco2e"] != 0.0:
                break

    # --- Scope 2 Electricity = grand_total - sum(scope_1_components) ---
    scope1_sum = (
        result["scope1_diesel_stationary_tco2e"]
        + result["scope1_diesel_mobile_tco2e"]
        + result["scope1_petrol_tco2e"]
        + result["scope1_hfhsd_ifo_tco2e"]
        + result["scope1_other_fuels_tco2e"]
    )
    result["scope2_electricity_tco2e"] = max(0.0, result["total_emissions_tco2e"] - scope1_sum)

    return result


def parse_workbook(workbook_path: str, port_id: str, fy: str = "FY24-25") -> HeadlineMetrics:
    cargo_val, cargo_ref = parse_cargo_total(workbook_path)
    m = parse_305_4(workbook_path)
    cell_refs = m["cell_refs"]
    cell_refs["cargo"] = cargo_ref
    return HeadlineMetrics(
        port_id=port_id, fy=fy, cargo_mt=cargo_val,
        total_emissions_tco2e=m["total_emissions_tco2e"],
        scope2_electricity_tco2e=m["scope2_electricity_tco2e"],
        scope1_diesel_stationary_tco2e=m["scope1_diesel_stationary_tco2e"],
        scope1_diesel_mobile_tco2e=m["scope1_diesel_mobile_tco2e"],
        scope1_petrol_tco2e=m["scope1_petrol_tco2e"],
        scope1_hfhsd_ifo_tco2e=m["scope1_hfhsd_ifo_tco2e"],
        scope1_other_fuels_tco2e=m["scope1_other_fuels_tco2e"],
        ghg_intensity_kg_per_mt=m["total_emissions_tco2e"] / cargo_val * 1000 if cargo_val else 0,
        cell_refs={k: asdict(v) for k, v in cell_refs.items()},
    )
