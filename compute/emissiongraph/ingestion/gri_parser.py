"""GRI sheet parsers for 302-1, 305-1, 305-2, 305-4.

Key rules from spec:
1. Cell-level provenance mandatory — every Measurement carries exact (sheet, cell).
2. Computed totals are NOT parsed as facts — recomputed downstream.
3. Annual-only sources (HFCs, Acetylene) carry period="annual".
4. Zero rows still emit Measurements for completeness.
5. Cross-sheet dedup: same fuel quantity appearing in 302-1/305-1/305-2
   is ONE underlying measurement. Dedup key: (port_id, fy, month, fuel_type, sub_type).
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

from emissiongraph.facts.schema import AmbiguousTotalWarning, CellRef, Measurement
from emissiongraph.ingestion.workbook_loader import get_sheet
from emissiongraph.ingestion.cargo_parser import (
    MONTH_NAMES,
    FY_MONTH_MAP,
    _col_letter,
    _fy_to_start_year,
    _month_to_period_value,
)

# Fuel type normalization map — maps common variations to canonical names
FUEL_NORMALIZE: dict[str, str] = {
    "electricity": "Electricity",
    "purchased electricity": "Electricity",
    "grid electricity": "Electricity",
    "diesel": "Diesel",
    "hsd": "Diesel",
    "high speed diesel": "Diesel",
    "diesel (stationary eqp)": "Diesel",
    "diesel (stationary eqp.)": "Diesel",
    "diesel (mobile eqp)": "Diesel",
    "diesel (mobile eqp.)": "Diesel",
    "petrol": "Petrol",
    "motor spirit": "Petrol",
    "petrol (mobile eqp)": "Petrol",
    "petrol (mobile eqp.)": "Petrol",
    "furnace oil": "Furnace Oil",
    "fo": "Furnace Oil",
    "lpg": "LPG",
    "coal": "Coal",
    "hfc": "HFC",
    "hfcs": "HFC",
    "r-410a": "HFC",
    "refrigerant": "HFC",
    "acetylene": "Acetylene",
    "biodiesel": "Biodiesel",
}

# Fuels that are annual-only (no monthly breakdown available)
ANNUAL_ONLY_FUELS = {"HFC", "Acetylene"}

# Fuels with sub_type disambiguation
DIESEL_SUBTYPES = {
    "diesel (stationary eqp)": "stationary",
    "diesel (stationary eqp.)": "stationary",
    "diesel (mobile eqp)": "mobile",
    "diesel (mobile eqp.)": "mobile",
}

PETROL_SUBTYPES = {
    "petrol (mobile eqp)": "mobile",
    "petrol (mobile eqp.)": "mobile",
}

# Unit detection
UNIT_HINTS = {
    "Electricity": "MWH",
    "Diesel": "KL",
    "Petrol": "KL",
    "Furnace Oil": "KL",
    "LPG": "T",
    "Coal": "T",
    "HFC": "Kg",
    "Acetylene": "Kg",
    "Biodiesel": "KL",
}


def _normalize_fuel(raw_label: str) -> tuple[str, str | None]:
    """Normalize a fuel label to (canonical_fuel_type, sub_type)."""
    lower = raw_label.strip().lower()

    # Check diesel subtypes first
    for key, sub in DIESEL_SUBTYPES.items():
        if lower == key:
            return "Diesel", sub

    # Check petrol subtypes
    for key, sub in PETROL_SUBTYPES.items():
        if lower == key:
            return "Petrol", sub

    # General normalization
    canonical = FUEL_NORMALIZE.get(lower)
    if canonical:
        return canonical, None

    # Fuzzy: check if any key is a substring
    for key, canonical in FUEL_NORMALIZE.items():
        if key in lower:
            return canonical, None

    # Unknown fuel — return as-is
    return raw_label.strip(), None


def _detect_unit(fuel_type: str, header_text: str | None = None) -> str:
    """Detect the measurement unit for a fuel type."""
    if header_text:
        h = header_text.upper()
        for unit in ["MWH", "KL", "MT", "T", "KG"]:
            if unit in h:
                return unit
    return UNIT_HINTS.get(fuel_type, "T")


def _find_month_columns(ws: Worksheet, max_scan_rows: int = 10) -> tuple[int | None, dict[str, int]]:
    """Scan the first rows to find the header row and month column mapping."""
    for row_idx in range(1, max_scan_rows + 1):
        month_cols: dict[str, int] = {}
        for col_idx in range(1, (ws.max_column or 30) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str):
                normalized = val.strip().lower()
                if normalized in MONTH_NAMES:
                    month_cols[normalized] = col_idx
        if len(month_cols) >= 6:  # at least half the months found
            return row_idx, month_cols
    return None, {}


def _find_annual_column(ws: Worksheet, header_row: int) -> int | None:
    """Find the 'Total' or 'Annual' column in the header row."""
    for col_idx in range(1, (ws.max_column or 30) + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and isinstance(val, str):
            lower = val.strip().lower()
            if lower in ("total", "annual", "yearly", "fy total"):
                return col_idx
    return None


def _is_total_row(label: str) -> bool:
    """Check if a row label indicates a computed total."""
    lower = label.strip().lower()
    return any(kw in lower for kw in [
        "total", "sub-total", "subtotal", "grand total",
        "scope 1", "scope 2", "scope 1+2",
    ])


def parse_energy_sheet(
    ws: Worksheet,
    port_id: str,
    fy: str,
    workbook_name: str,
    sheet_name: str = "302-1",
) -> tuple[list[Measurement], list[AmbiguousTotalWarning]]:
    """Parse GRI 302-1 energy consumption sheet.

    Returns (measurements, warnings).
    """
    measurements: list[Measurement] = []
    warnings: list[AmbiguousTotalWarning] = []

    header_row, month_cols = _find_month_columns(ws)
    if header_row is None:
        return measurements, warnings

    annual_col = _find_annual_column(ws, header_row)

    # Parse data rows below header
    for row_idx in range(header_row + 1, (ws.max_row or 100) + 1):
        label_cell = ws.cell(row=row_idx, column=1).value
        if label_cell is None:
            continue
        label = str(label_cell).strip()
        if not label:
            continue

        # Skip total rows — we verify them but don't store as facts
        if _is_total_row(label):
            # Verify total if annual column exists
            if annual_col:
                parsed_total_val = ws.cell(row=row_idx, column=annual_col).value
                if parsed_total_val is not None:
                    try:
                        parsed_total = float(parsed_total_val)
                        # Sum the monthly values we've seen for verification
                        row_sum = 0.0
                        for _, col_idx in month_cols.items():
                            v = ws.cell(row=row_idx, column=col_idx).value
                            if v is not None:
                                try:
                                    row_sum += float(v)
                                except (ValueError, TypeError):
                                    pass
                        if parsed_total != 0 and abs(row_sum - parsed_total) / abs(parsed_total) > 0.005:
                            warnings.append(AmbiguousTotalWarning(
                                sheet=sheet_name,
                                row_label=label,
                                parsed_total=parsed_total,
                                computed_total=row_sum,
                                pct_diff=abs(row_sum - parsed_total) / abs(parsed_total) * 100,
                            ))
                    except (ValueError, TypeError):
                        pass
            continue

        fuel_type, sub_type = _normalize_fuel(label)
        unit = _detect_unit(fuel_type)
        is_annual = fuel_type in ANNUAL_ONLY_FUELS

        if is_annual:
            # Annual-only: try annual column or sum of monthly
            annual_val = None
            if annual_col:
                v = ws.cell(row=row_idx, column=annual_col).value
                if v is not None:
                    try:
                        annual_val = float(v)
                    except (ValueError, TypeError):
                        pass
            if annual_val is None:
                # Sum monthly values
                annual_val = 0.0
                for _, col_idx in month_cols.items():
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is not None:
                        try:
                            annual_val += float(v)
                        except (ValueError, TypeError):
                            pass

            col_for_ref = annual_col or 1
            measurements.append(Measurement(
                port_id=port_id,
                fy=fy,
                period="annual",
                period_value=fy,
                fuel_type=fuel_type,
                sub_type=sub_type,
                measure="fugitive_release" if fuel_type == "HFC" else "consumption",
                quantity=annual_val,
                unit=unit,
                source_cell=CellRef(
                    workbook=workbook_name,
                    sheet=sheet_name,
                    cell=f"{_col_letter(col_for_ref)}{row_idx}",
                    row=row_idx,
                    col=col_for_ref,
                ),
            ))
        else:
            # Monthly: extract each month
            for month_name, col_idx in month_cols.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                # Zero rows still emit measurements per spec rule 4
                qty = 0.0
                if val is not None:
                    try:
                        qty = float(val)
                    except (ValueError, TypeError):
                        continue

                period_value = _month_to_period_value(month_name, fy)
                if not period_value:
                    continue

                col_letter = _col_letter(col_idx)
                measurements.append(Measurement(
                    port_id=port_id,
                    fy=fy,
                    period="monthly",
                    period_value=period_value,
                    fuel_type=fuel_type,
                    sub_type=sub_type,
                    measure="consumption",
                    quantity=qty,
                    unit=unit,
                    source_cell=CellRef(
                        workbook=workbook_name,
                        sheet=sheet_name,
                        cell=f"{col_letter}{row_idx}",
                        row=row_idx,
                        col=col_idx,
                    ),
                ))

    return measurements, warnings


def parse_emissions_sheet(
    ws: Worksheet,
    port_id: str,
    fy: str,
    workbook_name: str,
    sheet_name: str,  # "305-1" or "305-2"
    scope: str,  # "scope1" or "scope2"
) -> tuple[list[Measurement], list[AmbiguousTotalWarning]]:
    """Parse GRI 305-1 (Scope 1) or 305-2 (Scope 2) emissions sheets.

    These sheets share fuel-quantity columns with 302-1. Per spec, the same
    fuel quantity appearing in multiple sheets is ONE underlying measurement.
    We parse the emission values (tCO2e) here — fuel consumption is parsed from 302-1.
    """
    # For 305-1/305-2, the structure is similar to 302-1 but values are in tCO2e
    # We parse them as separate measurements with a different semantic meaning
    # The deduplication happens at the Measurement ID level via the deterministic UUID5
    return parse_energy_sheet(ws, port_id, fy, workbook_name, sheet_name)


def parse_intensity_sheet(
    ws: Worksheet,
    port_id: str,
    fy: str,
    workbook_name: str,
) -> tuple[list[Measurement], list[AmbiguousTotalWarning]]:
    """Parse GRI 305-4 intensity metrics sheet.

    305-4 contains combined intensity (tCO2e per MT, GJ per MT).
    These are computed values — we parse them for verification but mark as INFERRED.
    """
    measurements: list[Measurement] = []
    warnings: list[AmbiguousTotalWarning] = []

    header_row, month_cols = _find_month_columns(ws)
    if header_row is None:
        return measurements, warnings

    for row_idx in range(header_row + 1, (ws.max_row or 100) + 1):
        label_cell = ws.cell(row=row_idx, column=1).value
        if label_cell is None:
            continue
        label = str(label_cell).strip()
        if not label:
            continue

        lower = label.lower()
        # Detect intensity metric type
        if "co2" in lower or "emission" in lower:
            fuel_type = "EmissionIntensity"
            unit = "tCO2e/MT"
        elif "energy" in lower or "gj" in lower:
            fuel_type = "EnergyIntensity"
            unit = "GJ/MT"
        else:
            continue

        for month_name, col_idx in month_cols.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            try:
                qty = float(val)
            except (ValueError, TypeError):
                continue

            period_value = _month_to_period_value(month_name, fy)
            if not period_value:
                continue

            measurements.append(Measurement(
                port_id=port_id,
                fy=fy,
                period="monthly",
                period_value=period_value,
                fuel_type=fuel_type,
                sub_type=None,
                measure="consumption",
                quantity=qty,
                unit=unit,
                source_cell=CellRef(
                    workbook=workbook_name,
                    sheet="305-4",
                    cell=f"{_col_letter(col_idx)}{row_idx}",
                    row=row_idx,
                    col=col_idx,
                ),
                confidence="INFERRED",
            ))

    return measurements, warnings


def parse_workbook(
    wb: Workbook,
    port_id: str,
    fy: str,
    workbook_name: str,
) -> tuple[list[Measurement], list[AmbiguousTotalWarning]]:
    """Parse all GRI sheets from a workbook, deduplicating across sheets.

    Deduplication key: (port_id, fy, period_value, fuel_type, sub_type).
    The deterministic UUID5 on Measurement handles this automatically —
    two measurements with the same key produce the same ID, and dict keying
    deduplicates.
    """
    all_measurements: dict[str, Measurement] = {}  # id -> Measurement (dedup)
    all_warnings: list[AmbiguousTotalWarning] = []

    # Parse Cargo Handled
    from emissiongraph.ingestion.cargo_parser import parse_cargo_sheet
    from emissiongraph.ingestion.workbook_loader import get_sheet

    cargo_ws = get_sheet(wb, "Cargo Handled")
    if cargo_ws:
        for m in parse_cargo_sheet(cargo_ws, port_id, fy, workbook_name):
            all_measurements[m.id] = m

    # Parse 302-1 (Energy consumption) — primary source for fuel quantities
    energy_ws = get_sheet(wb, "302-1")
    if energy_ws:
        ms, ws_ = parse_energy_sheet(energy_ws, port_id, fy, workbook_name, "302-1")
        for m in ms:
            all_measurements[m.id] = m
        all_warnings.extend(ws_)

    # Parse 305-1 (Scope 1 emissions)
    scope1_ws = get_sheet(wb, "305-1")
    if scope1_ws:
        ms, ws_ = parse_emissions_sheet(scope1_ws, port_id, fy, workbook_name, "305-1", "scope1")
        for m in ms:
            # Only add if not already present from 302-1 (dedup via ID)
            if m.id not in all_measurements:
                all_measurements[m.id] = m
        all_warnings.extend(ws_)

    # Parse 305-2 (Scope 2 emissions)
    scope2_ws = get_sheet(wb, "305-2")
    if scope2_ws:
        ms, ws_ = parse_emissions_sheet(scope2_ws, port_id, fy, workbook_name, "305-2", "scope2")
        for m in ms:
            if m.id not in all_measurements:
                all_measurements[m.id] = m
        all_warnings.extend(ws_)

    # Parse 305-4 (Intensity — for verification only)
    intensity_ws = get_sheet(wb, "305-4")
    if intensity_ws:
        ms, ws_ = parse_intensity_sheet(intensity_ws, port_id, fy, workbook_name)
        for m in ms:
            all_measurements[m.id] = m
        all_warnings.extend(ws_)

    return list(all_measurements.values()), all_warnings
