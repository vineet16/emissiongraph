"""Workbook loader — opens an Excel file with openpyxl, validates expected sheets."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook

EXPECTED_SHEETS = {
    "Cargo Handled",
    "302-1",
    "305-1",
    "305-2",
    "305-4",
}


class WorkbookLoadError(Exception):
    pass


def load_workbook(path: str | Path) -> Workbook:
    """Load workbook with openpyxl (data_only=False to keep formulas visible)."""
    p = Path(path)
    if not p.exists():
        raise WorkbookLoadError(f"File not found: {p}")
    if not p.suffix.lower() in (".xlsx", ".xls"):
        raise WorkbookLoadError(f"Not an Excel file: {p}")

    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=False)
    return wb


def validate_workbook_sheets(wb: Workbook) -> list[str]:
    """Check that all expected GRI sheets are present. Returns list of missing sheet names."""
    present = set(wb.sheetnames)
    missing = []
    for expected in EXPECTED_SHEETS:
        # Allow case-insensitive + whitespace-trimmed matching
        found = any(
            s.strip().lower() == expected.lower() for s in present
        )
        if not found:
            missing.append(expected)
    return missing


def get_sheet(wb: Workbook, name: str):
    """Get a sheet by name, case-insensitive."""
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.lower():
            return wb[sn]
    return None
